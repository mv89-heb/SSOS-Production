"""
Phase 3.2A — Import Analysis Engine.

Read-only structural analysis of an uploaded workbook: how many sheets,
header shape, wide/tall orientation, per-column type guesses, detected
suppliers/units, and data-quality issues. Nothing here creates or updates
a Product, Supplier, or SupplierProductOffer, and nothing here is read by
OrderService — this is purely informational, for a human (or the future
Mapping Engine) to review before any mapping decision is made.

Design note: Phase 3.1's ImportRow only holds the ONE sheet chosen at
upload time, already flattened to plain strings — that's not enough to
answer "how many sheets does this workbook have" or "are any cells
merged/hidden/formulas". So this service re-opens the ORIGINAL file from
ImportSession.storage_path directly, independently of the staged rows.
"""
import os
import re
from collections import Counter

from app.repositories.import_session_repository import ImportSessionRepository
from app.repositories.import_analysis_repository import ImportAnalysisRepository
from app.repositories.supplier_repository import SupplierRepository
from app.models.import_analysis import FORMAT_WIDE, FORMAT_TALL, FORMAT_MIXED, FORMAT_UNKNOWN
from app.services.audit_service import AuditService
from app.utils.header_detection import detect_header


class ImportAnalysisError(Exception):
    """Raised when the original uploaded file can no longer be read for analysis."""


# --- Column-type keyword vocabulary (checked in this priority order — a
# column is only ever assigned ONE type, the first keyword group it
# matches) -------------------------------------------------------------
COLUMN_KEYWORDS = [
    ("BARCODE", ["ברקוד", "barcode", "אן.אס.אי", "ean"]),
    ("PRODUCT_CODE", ['מק"ט', "מקט", "קוד מוצר", "קוד פריט", "product code", "sku"]),
    # "לפני/אחרי מע\"מ" are distinct, explicitly named in spec — checked
    # before the generic VAT fallback below.
    ("PRICE_BEFORE_VAT", ['לפני מע"מ', "לפני מעמ", "before vat"]),
    ("PRICE_AFTER_VAT", ['אחרי מע"מ', "אחרי מעמ", 'כולל מע"מ', "כולל מעמ", "after vat"]),
    ("VAT", ['מע"מ', "מעמ", "vat"]),
    # "מחיר אחרי הנחה" / "מחיר מבצע" are PRICE values (per spec: "מחיר רגיל,
    # מחיר מבצע, לפני/אחרי מע\"מ" are all price categories) even though they
    # contain "הנחה"/"מבצע" — must be checked before the generic DISCOUNT
    # match below, or a compound header like this gets misclassified as a
    # bare discount rate instead of an actual price.
    ("PRICE", ["מחיר אחרי הנחה", "מחיר מבצע", "מחיר לאחר הנחה", "מחיר סופי"]),
    ("DISCOUNT", ["הנחה", "מבצע", "discount", "sale"]),
    ("CATEGORY", ["קטגור", "category", "סוג"]),
    ("SUPPLIER", ["ספק", "supplier", "יצרן"]),
    ("UNIT", ["יחיד", "unit", "אריזה"]),
    ("QUANTITY", ["כמות", "qty", "quantity"]),
    ("NOTES", ["הער", "notes", "comment", "תיאור נוסף"]),
    ("PRODUCT_NAME", ["מוצר", "שם פריט", "product", "item name", "תיאור"]),
    ("PRICE", ["מחיר", "price", "עלות", "cost", '\u20aa']),
    ("CODE", ["קוד"]),  # generic "קוד" alone, after more specific code types above
]

KNOWN_UNITS = {"קילו", 'ק"ג', "קג", "יחידה", "יח'", "יח", "ארגז", "קרטון", "חבילה", "ליטר", "גרם"}
CURRENCY_SYMBOLS = ["\u20aa", "$", "\u20ac"]  # ₪, $, €

_NUMERIC_RE = re.compile(r"^-?\d+(\.\d+)?$")


def _clean_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_numeric(value) -> bool:
    if isinstance(value, (int, float)):
        return True
    s = _clean_str(value)
    if not s:
        return False
    # Strip currency symbols/commas before checking — a value like "20.85 ₪"
    # is still "numeric" in intent even though it isn't a bare number.
    stripped = s
    for sym in CURRENCY_SYMBOLS:
        stripped = stripped.replace(sym, "")
    stripped = stripped.replace(",", "").strip()
    return bool(_NUMERIC_RE.match(stripped))


def _forward_fill(row: list, width: int) -> list:
    """Carries the last non-empty value forward across empty cells — this
    is what a merged header cell (or a label a person only bothered to
    write once) means: "this label applies to this column too." Used to
    recover a top-tier supplier-name label even though openpyxl/xlrd only
    report a value in the top-left cell of a merged range."""
    filled = []
    last = None
    for i in range(width):
        val = _clean_str(row[i]) if i < len(row) else ""
        if val:
            last = val
        filled.append(last)
    return filled


def _row_signature(row: list) -> dict:
    non_empty = [v for v in row if _clean_str(v) != ""]
    numeric = [v for v in non_empty if _looks_numeric(v)]
    return {
        "non_empty_count": len(non_empty),
        "numeric_ratio": (len(numeric) / len(non_empty)) if non_empty else 0.0,
    }


class WorkbookAnalyzer:
    """Format-agnostic detection logic. Each _analyze_*_sheet() function
    gathers a (rows, extras) pair in its own format's terms and hands off
    to this shared brain, so the actual heuristics exist in exactly one
    place regardless of file type."""

    def __init__(self, known_supplier_names: dict):
        # {normalized_name: (supplier_id, real_name)}
        self.known_supplier_names = known_supplier_names

    def analyze_sheet(
        self,
        sheet_name: str,
        sheet_index: int,
        is_hidden: bool,
        rows: list,  # list[list[Any]], raw values, untouched types
        merged_ranges: list,  # list of (min_row, max_row, min_col, max_col), 1-based, or [] if unknown
        hidden_rows: set,  # 1-based row numbers, or set() if unknown
        hidden_cols: set,  # 1-based col numbers, or set() if unknown
        formula_cells: int | None,  # count, or None if this format can't tell
    ) -> dict:
        row_count = len(rows)
        column_count = max((len(r) for r in rows), default=0)
        warnings = []

        if row_count == 0:
            return {
                "sheet_name": sheet_name, "sheet_index": sheet_index, "is_hidden": is_hidden,
                "row_count": 0, "column_count": 0, "header_row_index": None,
                "header_tier_count": None, "has_merged_header_cells": False,
                "detected_format": FORMAT_UNKNOWN, "format_reason": "Sheet is empty.",
                "columns": [], "detected_suppliers": [], "detected_units": [],
                "data_quality": {}, "warnings": ["Sheet has no rows at all."],
            }

        header_start, tier_count, header_reason = self._detect_header(rows)
        warnings.append(header_reason)

        has_merged_header = any(
            min_r <= header_start + tier_count - 1 and max_r >= header_start
            for (min_r, max_r, _min_c, _max_c) in merged_ranges
        )

        headers = self._build_header_labels(rows, header_start, tier_count)
        raw_headers = self._raw_header_labels(rows, header_start, tier_count)
        group_labels = self._build_group_labels(rows, header_start, tier_count)
        data_rows = rows[header_start + tier_count - 1:]  # 0-based slice, after the header block

        columns = self._detect_columns(headers, data_rows, group_labels)
        detected_suppliers = self._detect_suppliers(columns)
        detected_units = self._detect_units(data_rows, columns)

        detected_format, format_reason = self._detect_orientation(columns, detected_suppliers)

        data_quality = self._detect_data_quality(
            headers, raw_headers, data_rows, columns, merged_ranges, hidden_rows, hidden_cols, formula_cells
        )

        return {
            "sheet_name": sheet_name,
            "sheet_index": sheet_index,
            "is_hidden": is_hidden,
            "row_count": row_count,
            "column_count": column_count,
            "header_row_index": header_start,
            "header_tier_count": tier_count,
            "has_merged_header_cells": has_merged_header,
            "detected_format": detected_format,
            "format_reason": format_reason,
            "columns": columns,
            "detected_suppliers": detected_suppliers,
            "detected_units": detected_units,
            "data_quality": data_quality,
            "warnings": warnings,
        }

    # ------------------------------------------------------------------
    # Header detection
    # ------------------------------------------------------------------
    @staticmethod
    def _detect_header(rows: list, max_tiers: int = 3) -> tuple:
        """Delegates to app.utils.header_detection.detect_header — the
        shared single source of truth also used by ImportService's
        Staging step (see that module's docstring for why this now lives
        in one place instead of two independent copies). Kept as a
        wrapper here, rather than inlining the import at every call site
        in this file, so nothing else in WorkbookAnalyzer has to change."""
        return detect_header(rows, max_tiers)

    @staticmethod
    def _build_header_labels(rows: list, header_start_1based: int, tier_count: int) -> list:
        """Uses only the LAST row of a multi-tier header block for column
        labels (closest to the actual data, e.g. "יחידה"/"מחיר" rather than
        a sparse supplier-name row) — but every tier's text is preserved in
        `warnings`/`format_reason` for the human to see, nothing is thrown
        away silently. Duplicate labels get a " (2)", " (3)"... suffix,
        same rule as Phase 3.1's raw capture, so no column is ever dropped."""
        label_row_idx = header_start_1based - 1 + tier_count - 1
        label_row = rows[label_row_idx] if label_row_idx < len(rows) else []
        width = max((len(r) for r in rows), default=0)

        labels = []
        seen = Counter()
        for i in range(width):
            cell = label_row[i] if i < len(label_row) else None
            name = _clean_str(cell) or f"עמודה {i + 1}"
            seen[name] += 1
            if seen[name] > 1:
                name = f"{name} ({seen[name]})"
            labels.append(name)
        return labels

    @staticmethod
    def _raw_header_labels(rows: list, header_start_1based: int, tier_count: int) -> list:
        """Same label row as _build_header_labels, but WITHOUT the " (2)"
        disambiguation suffix and without inventing "עמודה N" placeholders
        for blanks — used only for duplicate-header detection, which needs
        to see the actual collision before it gets resolved away."""
        label_row_idx = header_start_1based - 1 + tier_count - 1
        label_row = rows[label_row_idx] if label_row_idx < len(rows) else []
        width = max((len(r) for r in rows), default=0)
        return [_clean_str(label_row[i]) if i < len(label_row) else "" for i in range(width)]

    @staticmethod
    def _build_group_labels(rows: list, header_start_1based: int, tier_count: int) -> list:
        """For a multi-tier header, the TOP tier often carries the real
        group identity (e.g. a supplier name merged across several
        sub-columns) that _build_header_labels intentionally doesn't use
        for column names, to avoid sparse/duplicate labels. This recovers
        that identity per column via forward-fill, so e.g. a "לפני מע\"מ"
        column can still be attributed to "גידרון" even though the column
        NAME is just "לפני מע\"מ". Returns [] for a single-tier header —
        there's no separate group tier to recover anything from."""
        if tier_count <= 1:
            return []
        top_row = rows[header_start_1based - 1] if header_start_1based - 1 < len(rows) else []
        width = max((len(r) for r in rows), default=0)
        return _forward_fill(top_row, width)

    # ------------------------------------------------------------------
    # Column type detection
    # ------------------------------------------------------------------
    def _detect_columns(self, headers: list, data_rows: list, group_labels: list) -> list:
        columns = []
        for idx, header in enumerate(headers):
            samples = []
            for row in data_rows:
                if idx < len(row) and _clean_str(row[idx]) != "":
                    samples.append(row[idx])
                if len(samples) >= 15:
                    break

            group_label = group_labels[idx] if idx < len(group_labels) else None
            col_type, confidence, reason = self._classify_column(header, samples)
            columns.append({
                "index": idx,
                "header": header,
                "group_label": group_label,
                "detected_type": col_type,
                "confidence": confidence,
                "reason": reason,
                "sample_values": [_clean_str(v) for v in samples[:5]],
            })
        return columns

    def _classify_column(self, header: str, samples: list) -> tuple:
        header_norm = header.strip().lower()

        # Strong signal: header text matches an existing supplier's name in
        # this tenant's real catalog data.
        if header_norm in self.known_supplier_names:
            supplier_id, real_name = self.known_supplier_names[header_norm]
            return "SUPPLIER", "high", f'Header matches existing supplier "{real_name}".'

        for col_type, keywords in COLUMN_KEYWORDS:
            if any(kw in header_norm for kw in keywords):
                shape_note = self._describe_sample_shape(samples)
                return col_type, "medium", f'Header contains "{[kw for kw in keywords if kw in header_norm][0]}". {shape_note}'

        # No keyword match — fall back to sample-value shape alone. Require
        # a few samples before guessing anything: a phantom "used range"
        # column with just one stray leftover value (confirmed against a
        # real uploaded file) must not be confidently typed off a sample
        # size of 1 — that pollutes orientation detection downstream.
        if len(samples) >= 3:
            numeric_ratio = sum(1 for s in samples if _looks_numeric(s)) / len(samples)
            if numeric_ratio >= 0.8:
                return "PRICE", "low", f"No header keyword matched, but {numeric_ratio:.0%} of sampled values look numeric — guessing PRICE."
            unit_ratio = sum(1 for s in samples if _clean_str(s) in KNOWN_UNITS) / len(samples)
            if unit_ratio >= 0.5:
                return "UNIT", "low", f"No header keyword matched, but {unit_ratio:.0%} of sampled values are known unit words — guessing UNIT."

        return "UNKNOWN", "none", "No header keyword matched and sample values gave no reliable signal."

    @staticmethod
    def _describe_sample_shape(samples: list) -> str:
        if not samples:
            return "No sample data available to corroborate."
        numeric_ratio = sum(1 for s in samples if _looks_numeric(s)) / len(samples)
        return f"{numeric_ratio:.0%} of sampled values look numeric."

    # ------------------------------------------------------------------
    # Supplier / unit detection
    # ------------------------------------------------------------------
    def _detect_suppliers(self, columns: list) -> list:
        """A column can indicate a supplier two ways: its own header text
        matches a known supplier (single-tier headers, e.g. a plain
        "גידרון" column), or its recovered group_label does (multi-tier
        headers, e.g. "לפני מע\"מ" sitting under a merged "גידרון" cell).
        Both are reported; a header-level SUPPLIER match is also always
        surfaced even without a catalog match, at lower confidence."""
        found = []
        seen_columns = set()

        for col in columns:
            if col["detected_type"] == "SUPPLIER" and col["index"] not in seen_columns:
                header_norm = col["header"].strip().lower()
                match = self.known_supplier_names.get(header_norm)
                found.append({
                    "column_index": col["index"],
                    "header": col["header"],
                    "source": "header",
                    "matched_supplier_id": match[0] if match else None,
                    "matched_supplier_name": match[1] if match else None,
                })
                seen_columns.add(col["index"])

        for col in columns:
            if col["index"] in seen_columns or not col.get("group_label"):
                continue
            group_norm = col["group_label"].strip().lower()
            match = self.known_supplier_names.get(group_norm)
            if match:
                found.append({
                    "column_index": col["index"],
                    "header": col["group_label"],
                    "source": "group_label",
                    "matched_supplier_id": match[0],
                    "matched_supplier_name": match[1],
                })
                seen_columns.add(col["index"])

        return found

    @staticmethod
    def _detect_units(data_rows: list, columns: list) -> list:
        found = set()
        unit_column_indices = [c["index"] for c in columns if c["detected_type"] == "UNIT"]
        search_indices = unit_column_indices or range(len(columns))
        for row in data_rows:
            for idx in search_indices:
                if idx < len(row):
                    val = _clean_str(row[idx])
                    if val in KNOWN_UNITS:
                        found.add(val)
        return sorted(found)

    # ------------------------------------------------------------------
    # Orientation (wide/tall/mixed)
    # ------------------------------------------------------------------
    @staticmethod
    def _detect_orientation(columns: list, detected_suppliers: list) -> tuple:
        price_like_types = ("PRICE", "VAT", "PRICE_BEFORE_VAT", "PRICE_AFTER_VAT", "DISCOUNT")
        reliable = lambda c: c["confidence"] in ("high", "medium")
        price_columns = [c for c in columns if c["detected_type"] == "PRICE" and reliable(c)]
        price_like_columns = [c for c in columns if c["detected_type"] in price_like_types and reliable(c)]

        # Distinct group labels among price-like columns — this is what
        # catches a merged supplier-name row sitting above per-supplier
        # price sub-columns (e.g. גידרון/עמית/ווגשל), even when none of
        # those sub-columns' own headers say "supplier" and none matched
        # an existing catalog Supplier by name.
        group_label_suppliers = sorted({
            c["group_label"] for c in price_like_columns if c.get("group_label")
        })

        distinct_supplier_signal = max(len(detected_suppliers), len(group_label_suppliers))

        if distinct_supplier_signal >= 2:
            if len(group_label_suppliers) >= len(detected_suppliers):
                names = ", ".join(group_label_suppliers)
                source = "merged header group labels"
            else:
                names = ", ".join(s["header"] for s in detected_suppliers)
                source = "header/catalog match"
            return FORMAT_WIDE, (
                f"Found {distinct_supplier_signal} distinct supplier-like columns ({names}), "
                f"detected via {source} — each row appears to compare the same product across "
                "multiple suppliers."
            )
        if len(price_columns) == 1 and distinct_supplier_signal <= 1:
            return FORMAT_TALL, (
                "Found exactly one price column and at most one supplier column — "
                "each row appears to be one product from one supplier."
            )
        if len(price_columns) == 0 and not price_like_columns:
            return FORMAT_UNKNOWN, "No column was confidently classified as a price column — cannot determine orientation."
        return FORMAT_MIXED, (
            f"Found {len(price_columns)} price column(s), {len(price_like_columns)} price-like "
            f"(price/VAT/discount) column(s), and {distinct_supplier_signal} supplier-like column(s) — "
            "the pattern doesn't clearly match either a single-supplier or multi-supplier layout."
        )

    # ------------------------------------------------------------------
    # Data quality
    # ------------------------------------------------------------------
    def _detect_data_quality(
        self, headers, raw_headers, data_rows, columns, merged_ranges, hidden_rows, hidden_cols, formula_cells
    ) -> dict:
        empty_row_count = sum(1 for row in data_rows if _row_signature(row)["non_empty_count"] == 0)

        empty_columns = []
        for idx, header in enumerate(headers):
            if all(idx >= len(row) or _clean_str(row[idx]) == "" for row in data_rows):
                empty_columns.append(header)

        # Columns whose header cell itself was blank in the source file —
        # they got a "עמודה N" placeholder name in `headers`, but the
        # underlying cell was empty/missing. Reported by final (possibly
        # disambiguated) label so it's identifiable in the columns list.
        empty_header_columns = [
            headers[i] for i, raw in enumerate(raw_headers) if not raw and i < len(headers)
        ]

        # Duplicates must be checked BEFORE the " (2)"/"(3)" disambiguation
        # in _build_header_labels resolves the collision away — checking
        # the already-disambiguated `headers` here would never find one.
        raw_counts = Counter(h for h in raw_headers if h)
        duplicate_headers = sorted(h for h, n in raw_counts.items() if n > 1)

        product_col = next((c for c in columns if c["detected_type"] == "PRODUCT_NAME"), None)
        duplicate_products = []
        if product_col:
            values = [
                _clean_str(row[product_col["index"]])
                for row in data_rows
                if product_col["index"] < len(row) and _clean_str(row[product_col["index"]])
            ]
            counts = Counter(values)
            duplicate_products = sorted(v for v, n in counts.items() if n > 1)[:20]

        mixed_type_columns = []
        invalid_numeric_examples = {}
        for col in columns:
            idx = col["index"]
            values = [row[idx] for row in data_rows if idx < len(row) and _clean_str(row[idx]) != ""]
            if not values:
                continue
            numeric_flags = [_looks_numeric(v) for v in values]
            if any(numeric_flags) and not all(numeric_flags):
                mixed_type_columns.append(col["header"])
            if col["detected_type"] in ("PRICE", "QUANTITY", "VAT", "PRICE_BEFORE_VAT", "PRICE_AFTER_VAT", "DISCOUNT"):
                bad = [_clean_str(v) for v, is_num in zip(values, numeric_flags) if not is_num]
                if bad:
                    invalid_numeric_examples[col["header"]] = bad[:5]

        currency_columns = []
        for col in columns:
            idx = col["index"]
            for row in data_rows[:50]:
                if idx < len(row):
                    val = _clean_str(row[idx])
                    if any(sym in val for sym in CURRENCY_SYMBOLS):
                        currency_columns.append(col["header"])
                        break

        extra_whitespace_count = sum(
            1 for row in data_rows for v in row
            if isinstance(v, str) and v != v.strip() and v.strip() != ""
        )

        encoding_issue_count = sum(
            1 for row in data_rows for v in row
            if isinstance(v, str) and "\ufffd" in v
        )

        return {
            "empty_row_count": empty_row_count,
            "empty_columns": empty_columns,
            "empty_header_columns": empty_header_columns,
            "merged_cell_count": len(merged_ranges),
            "hidden_row_count": len(hidden_rows),
            "hidden_column_count": len(hidden_cols),
            "duplicate_headers": duplicate_headers,
            "duplicate_products": duplicate_products,
            "mixed_type_columns": mixed_type_columns,
            "formula_cell_count": formula_cells,  # None means "not detectable for this file format"
            "invalid_numeric_examples": invalid_numeric_examples,
            "currency_symbol_columns": sorted(set(currency_columns)),
            "extra_whitespace_cell_count": extra_whitespace_count,
            "encoding_issue_cell_count": encoding_issue_count,
        }


class ImportAnalysisService:
    def __init__(self, tenant_id: int, user_id: int):
        self.tenant_id = tenant_id
        self.user_id = user_id
        self.session_repo = ImportSessionRepository(tenant_id)
        self.analysis_repo = ImportAnalysisRepository(tenant_id)
        self.supplier_repo = SupplierRepository(tenant_id)

    def get_analysis(self, session_id: int) -> list:
        self.session_repo.get_by_id_or_404(session_id)  # tenant ownership check
        return self.analysis_repo.get_by_session(session_id)

    def analyze(self, session_id: int) -> list:
        """Re-opens the ORIGINAL uploaded file and analyzes every sheet.
        Read-only against the file; only ever writes import_analyses rows
        and the two workbook-summary columns on this ImportSession. Safe
        to call more than once — previous findings for this session are
        replaced, not accumulated."""
        session = self.session_repo.get_by_id_or_404(session_id)
        if not session.storage_path or not os.path.exists(session.storage_path):
            raise ImportAnalysisError("The originally uploaded file is no longer available on disk.")

        known_suppliers = {
            s.name.strip().lower(): (s.id, s.name) for s in self.supplier_repo.get_active()
        }
        analyzer = WorkbookAnalyzer(known_suppliers)

        ext = os.path.splitext(session.storage_path)[1].lower()
        if ext == ".xlsx":
            sheet_reports, sheet_names, active_sheet = self._analyze_xlsx(session.storage_path, analyzer)
        elif ext == ".xls":
            sheet_reports, sheet_names, active_sheet = self._analyze_xls(session.storage_path, analyzer)
        elif ext == ".csv":
            sheet_reports, sheet_names, active_sheet = self._analyze_csv(session.storage_path, analyzer)
        else:
            raise ImportAnalysisError(f"Unsupported file extension for analysis: {ext}")

        self.analysis_repo.delete_by_session(session_id)
        analysis_rows = []
        for report in sheet_reports:
            row = self.analysis_repo.model(
                tenant_id=self.tenant_id,
                import_session_id=session_id,
                **report,
            )
            self.analysis_repo.add(row)
            analysis_rows.append(row)

        session.workbook_sheet_names = sheet_names
        session.workbook_sheet_count = len(sheet_names)
        session.workbook_active_sheet = active_sheet

        AuditService.log_event(
            self.tenant_id, self.user_id, "import.analyzed",
            f"Analyzed {session.filename} ({len(sheet_names)} sheet(s))",
            {"import_session_id": session_id, "sheet_count": len(sheet_names)},
        )
        return analysis_rows

    # ------------------------------------------------------------------
    # Format-specific readers — each gathers (rows, merged_ranges,
    # hidden_rows, hidden_cols, formula_cells) in its own format's terms
    # and hands off to WorkbookAnalyzer.analyze_sheet, which holds all the
    # actual detection logic.
    # ------------------------------------------------------------------

    @staticmethod
    def _analyze_xlsx(path: str, analyzer: WorkbookAnalyzer):
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportAnalysisError("openpyxl is required to analyze .xlsx files") from exc

        # read_only=False is required here (unlike Phase 3.1's streaming
        # parse) — read-only worksheets have no merged_cells/row_dimensions/
        # column_dimensions attributes at all. data_only=False is required
        # to see formula text rather than only the last cached value.
        try:
            wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
        except Exception as exc:
            raise ImportAnalysisError(f"Could not open .xlsx file for analysis: {exc}") from exc

        reports = []
        for sheet_index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            is_hidden = ws.sheet_state != "visible"

            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            # Trim fully-empty trailing rows/columns from Excel's inflated
            # "used range" (a common artifact — see Phase 3.1 finding).
            rows = ImportAnalysisService._trim_trailing_empty(rows)

            merged_ranges = [
                (r.min_row, r.max_row, r.min_col, r.max_col) for r in ws.merged_cells.ranges
            ]
            hidden_rows = {
                i for i in range(1, len(rows) + 1)
                if ws.row_dimensions[i].hidden
            }
            hidden_cols = {
                i for i in range(1, (max((len(r) for r in rows), default=0)) + 1)
                if ws.column_dimensions[openpyxl.utils.get_column_letter(i)].hidden
            }
            formula_count = sum(
                1 for row in ws.iter_rows()
                for cell in row
                if isinstance(cell.value, str) and cell.value.startswith("=")
            )

            report = analyzer.analyze_sheet(
                sheet_name, sheet_index, is_hidden, rows,
                merged_ranges, hidden_rows, hidden_cols, formula_count,
            )
            reports.append(report)

        return reports, list(wb.sheetnames), wb.active.title if wb.active else None

    @staticmethod
    def _analyze_xls(path: str, analyzer: WorkbookAnalyzer):
        try:
            import xlrd
        except ImportError as exc:
            raise ImportAnalysisError("xlrd is required to analyze legacy .xls files") from exc

        try:
            wb = xlrd.open_workbook(path, formatting_info=True)
        except Exception:
            # formatting_info=True (needed for merged-cell/hidden detection)
            # fails on some legacy files (e.g. >65536 rows or odd XF
            # records) — fall back to a plain open so analysis can still
            # run, just without those two specific signals.
            try:
                wb = xlrd.open_workbook(path)
            except Exception as exc:
                raise ImportAnalysisError(f"Could not open .xls file for analysis: {exc}") from exc

        reports = []
        sheet_names = wb.sheet_names()
        for sheet_index, sheet_name in enumerate(sheet_names):
            sheet = wb.sheet_by_index(sheet_index)
            is_hidden = getattr(sheet, "visibility", 0) != 0

            rows = [
                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ]
            rows = ImportAnalysisService._trim_trailing_empty(rows)

            merged_ranges = []
            try:
                for (rlo, rhi, clo, chi) in getattr(sheet, "merged_cells", []):
                    merged_ranges.append((rlo + 1, rhi, clo + 1, chi))  # xlrd is 0-based, half-open
            except Exception:
                pass

            hidden_rows = set()
            hidden_cols = set()
            try:
                for r in range(sheet.nrows):
                    if sheet.rowinfo_map.get(r) and sheet.rowinfo_map[r].hidden:
                        hidden_rows.add(r + 1)
                for c in range(sheet.ncols):
                    if sheet.colinfo_map.get(c) and sheet.colinfo_map[c].hidden:
                        hidden_cols.add(c + 1)
            except Exception:
                pass

            # xlrd does not expose formula text for legacy .xls — always
            # reports the cached result, so this is honestly "unknown"
            # rather than guessed as zero.
            formula_count = None

            report = analyzer.analyze_sheet(
                sheet_name, sheet_index, is_hidden, rows,
                merged_ranges, hidden_rows, hidden_cols, formula_count,
            )
            reports.append(report)

        return reports, list(sheet_names), None  # xlrd doesn't expose the active/selected sheet

    @staticmethod
    def _analyze_csv(path: str, analyzer: WorkbookAnalyzer):
        import csv
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))
        except UnicodeDecodeError as exc:
            raise ImportAnalysisError(f"Could not read CSV as UTF-8: {exc}") from exc
        except OSError as exc:
            raise ImportAnalysisError(f"Could not open .csv file for analysis: {exc}") from exc

        rows = ImportAnalysisService._trim_trailing_empty(rows)
        # CSV has no concept of sheets, merged cells, hidden rows/cols, or
        # formulas — those signals are simply unavailable for this format.
        report = analyzer.analyze_sheet(
            "CSV", 0, False, rows, merged_ranges=[], hidden_rows=set(), hidden_cols=set(), formula_cells=None,
        )
        return [report], ["CSV"], None

    @staticmethod
    def _trim_trailing_empty(rows: list) -> list:
        """Drops fully-empty trailing rows (Excel's inflated 'used range'
        artifact, confirmed against a real uploaded file in Phase 3.1).
        Never removes a row/column that has any real data, and never
        touches anything in the middle of the sheet."""
        while rows and _row_signature(rows[-1])["non_empty_count"] == 0:
            rows.pop()
        return rows
