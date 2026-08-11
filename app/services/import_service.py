import os
import csv

from app.repositories.import_session_repository import ImportSessionRepository
from app.repositories.import_row_repository import ImportRowRepository
from app.repositories.supplier_repository import SupplierRepository
from app.models.import_session import ImportSession, STATUS_UPLOADED, STATUS_FAILED
from app.services.audit_service import AuditService
from app.utils.header_detection import detect_header


class ImportParseError(Exception):
    """Raised when a file can't be parsed at all (bad format, corrupt,
    unsupported extension, invalid sheet selection, or zero usable rows)."""


class ImportService:
    """Import staging layer for supplier Excel/CSV files."""

    def __init__(self, tenant_id: int, user_id: int):
        self.tenant_id = tenant_id
        self.user_id = user_id
        self.session_repo = ImportSessionRepository(tenant_id)
        self.row_repo = ImportRowRepository(tenant_id)
        self.supplier_repo = SupplierRepository(tenant_id)

    def list_sessions(self, limit: int = 50):
        return self.session_repo.list_recent(limit=limit)

    def get_session(self, session_id: int) -> ImportSession:
        return self.session_repo.get_by_id_or_404(session_id)

    def get_session_rows(self, session_id: int, limit: int = 100, offset: int = 0):
        self.session_repo.get_by_id_or_404(session_id)
        return self.row_repo.get_by_session(session_id, limit=limit, offset=offset)

    def create_session_and_parse(
        self, filename: str, storage_path: str, supplier_id: int = None, sheet_name: str = None
    ) -> ImportSession:
        if supplier_id is not None:
            self.supplier_repo.get_by_id_or_404(supplier_id)

        session = self.session_repo.model(
            tenant_id=self.tenant_id,
            filename=filename,
            storage_path=storage_path,
            supplier_id=supplier_id,
            uploaded_by=self.user_id,
            status=STATUS_UPLOADED,
        )
        self.session_repo.add(session)

        try:
            headers, data_rows, data_rows_values, resolved_sheet_name = self._parse_file(
                storage_path, sheet_name
            )
            if not data_rows:
                raise ImportParseError("No data rows found in file")
        except ImportParseError as exc:
            session.status = STATUS_FAILED
            session.error_message = str(exc)
            AuditService.log_event(
                self.tenant_id, self.user_id, "import.failed",
                f"Failed to parse {filename}: {exc}",
                {"import_session_id": session.id},
            )
            return session

        row_entities = [
            self.row_repo.model(
                tenant_id=self.tenant_id,
                import_session_id=session.id,
                row_number=i + 1,
                raw_data=row,
                raw_values=values,
            )
            for i, (row, values) in enumerate(zip(data_rows, data_rows_values))
        ]
        self.row_repo.bulk_add(row_entities)

        session.column_headers = headers
        session.row_count = len(data_rows)
        session.staged_sheet_name = resolved_sheet_name

        AuditService.log_event(
            self.tenant_id, self.user_id, "import.session_created",
            f"Uploaded {filename} ({len(data_rows)} rows)",
            {"import_session_id": session.id, "row_count": len(data_rows)},
        )
        return session

    @staticmethod
    def _parse_file(path: str, sheet_name: str = None):
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx":
            return ImportService._parse_xlsx(path, sheet_name)
        if ext == ".xls":
            return ImportService._parse_xls(path, sheet_name)
        if ext == ".csv":
            if sheet_name:
                raise ImportParseError("CSV files do not support sheet selection")
            return ImportService._parse_csv(path)
        raise ImportParseError(f"Unsupported file extension: {ext}")

    @staticmethod
    def _parse_xlsx(path: str, sheet_name: str = None):
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportParseError("openpyxl is required to read .xlsx files") from exc

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception as exc:
            raise ImportParseError(f"Could not open .xlsx file: {exc}") from exc

        try:
            sheet_names = list(wb.sheetnames)
            if not sheet_names:
                raise ImportParseError("The .xlsx workbook contains no worksheets")
            if sheet_name is not None and sheet_name not in sheet_names:
                raise ImportParseError(
                    f"Worksheet '{sheet_name}' was not found. Available worksheets: {', '.join(sheet_names)}"
                )
            resolved_name = sheet_name or sheet_names[0]
            ws = wb[resolved_name]
            headers, data_rows, data_rows_values = ImportService._rows_to_dicts(
                ws.iter_rows(values_only=True)
            )
            return headers, data_rows, data_rows_values, resolved_name
        except ImportParseError:
            raise
        except Exception as exc:
            raise ImportParseError(f"Could not parse worksheet '{sheet_name or sheet_names[0]}': {exc}") from exc
        finally:
            try:
                wb.close()
            except Exception:
                pass

    @staticmethod
    def _parse_xls(path: str, sheet_name: str = None):
        try:
            import xlrd
        except ImportError as exc:
            raise ImportParseError("xlrd is required to read legacy .xls files") from exc

        try:
            wb = xlrd.open_workbook(path)
        except Exception as exc:
            raise ImportParseError(f"Could not open .xls file: {exc}") from exc

        try:
            sheet_names = wb.sheet_names()
            if not sheet_names:
                raise ImportParseError("The .xls workbook contains no worksheets")
            if sheet_name is not None and sheet_name not in sheet_names:
                raise ImportParseError(
                    f"Worksheet '{sheet_name}' was not found. Available worksheets: {', '.join(sheet_names)}"
                )
            resolved_name = sheet_name or sheet_names[0]
            ws = wb.sheet_by_name(resolved_name)

            def _row_gen():
                for r in range(ws.nrows):
                    yield tuple(ws.cell_value(r, c) for c in range(ws.ncols))

            headers, data_rows, data_rows_values = ImportService._rows_to_dicts(_row_gen())
            return headers, data_rows, data_rows_values, resolved_name
        except ImportParseError:
            raise
        except Exception as exc:
            raise ImportParseError(f"Could not parse worksheet '{sheet_name or sheet_names[0]}': {exc}") from exc
        finally:
            try:
                wb.release_resources()
            except Exception:
                pass

    @staticmethod
    def _parse_csv(path: str):
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))
        except UnicodeDecodeError as exc:
            raise ImportParseError(f"Could not read CSV as UTF-8: {exc}") from exc
        except OSError as exc:
            raise ImportParseError(f"Could not open .csv file: {exc}") from exc

        headers, data_rows, data_rows_values = ImportService._rows_to_dicts(iter(rows))
        return headers, data_rows, data_rows_values, "CSV"

    @staticmethod
    def _rows_to_dicts(rows_iter):
        rows_list = list(rows_iter)

        if not any(row and any(cell not in (None, "") for cell in row) for row in rows_list):
            return [], [], []

        header_start_1based, tier_count, _reason = detect_header(rows_list)
        header_idx = header_start_1based - 1
        label_row_idx = header_idx + tier_count - 1
        label_row = rows_list[label_row_idx] if label_row_idx < len(rows_list) else []

        headers = []
        seen = {}
        for i, cell in enumerate(label_row):
            name = str(cell).strip() if cell not in (None, "") else f"עמודה {i + 1}"
            if name in seen:
                seen[name] += 1
                name = f"{name} ({seen[name]})"
            else:
                seen[name] = 1
            headers.append(name)

        data_rows = []
        data_rows_values = []
        for row in rows_list[header_idx + tier_count:]:
            if not row or all(cell in (None, "") for cell in row):
                continue
            row_dict = {}
            row_values = []
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else None
                str_value = "" if value is None else str(value)
                row_dict[header] = str_value
                row_values.append(str_value)
            data_rows.append(row_dict)
            data_rows_values.append(row_values)

        return headers, data_rows, data_rows_values
