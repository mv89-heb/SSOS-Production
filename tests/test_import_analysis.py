import io
import os
import openpyxl


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REAL_FILE_PATH = "/mnt/user-data/uploads/1784449037643_מחירים_-_ספקי_מזון.xls"


def _xlsx_bytes(rows, merges=None, hidden_sheet=False, second_sheet_rows=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in rows:
        ws.append(row)
    for merge_range in (merges or []):
        ws.merge_cells(merge_range)
    if second_sheet_rows is not None:
        ws2 = wb.create_sheet("Sheet2")
        for row in second_sheet_rows:
            ws2.append(row)
        if hidden_sheet:
            ws2.sheet_state = "hidden"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _upload(client, data, filename="test.xlsx", mime=XLSX_MIME):
    return client.post(
        "/api/imports/upload",
        data={"file": (io.BytesIO(data), filename, mime)},
        content_type="multipart/form-data",
    )


def _upload_and_analyze(client, data, **kwargs):
    session_id = _upload(client, data, **kwargs).get_json()["session"]["id"]
    resp = client.post(f"/api/imports/{session_id}/analyze")
    return session_id, resp


# --- Tall format -----------------------------------------------------------

def test_tall_format_detected(logged_in_client_a):
    data = _xlsx_bytes([
        ["מוצר", "יחידה", "מחיר"],
        ["חלה קלועה", "יחידה", "6.54"],
        ["פיתות", "חבילה", "8.0"],
    ])
    session_id, resp = _upload_and_analyze(logged_in_client_a, data)
    assert resp.status_code == 200
    sheet = resp.get_json()["analysis"][0]
    assert sheet["detected_format"] == "TALL"
    assert sheet["header_tier_count"] == 1
    types = {c["header"]: c["detected_type"] for c in sheet["columns"]}
    assert types["מוצר"] == "PRODUCT_NAME"
    assert types["מחיר"] == "PRICE"
    assert types["יחידה"] == "UNIT"


# --- Wide format + two-tier header + merged cells ---------------------------

def test_wide_format_two_tier_merged_header_detected(logged_in_client_a):
    logged_in_client_a.post("/api/catalog/suppliers", json={"name": "עמית"})
    data = _xlsx_bytes(
        [
            [None, "גידרון", None, None, "עמית", "ווגשל"],
            ["מוצר", "יחידה", 'לפני מע"מ', "הנחה", 'לפני מע"מ', 'לפני מע"מ'],
            ["בורקס גבינה", "קילו", 20.85, 14.6, 12.8, 14.64],
        ],
        merges=["B1:D1"],
    )
    session_id, resp = _upload_and_analyze(logged_in_client_a, data)
    assert resp.status_code == 200
    sheet = resp.get_json()["analysis"][0]

    assert sheet["header_tier_count"] == 2
    assert sheet["has_merged_header_cells"] is True
    assert sheet["detected_format"] == "WIDE"

    group_labels = {c["header"]: c["group_label"] for c in sheet["columns"]}
    assert group_labels['לפני מע"מ'] == "גידרון"
    assert group_labels['לפני מע"מ (2)'] == "עמית"
    assert group_labels['לפני מע"מ (3)'] == "ווגשל"

    supplier_names = {s["header"] for s in sheet["detected_suppliers"]}
    assert "עמית" in supplier_names
    matched = next(s for s in sheet["detected_suppliers"] if s["header"] == "עמית")
    assert matched["matched_supplier_id"] is not None


# --- Multiple sheets + hidden sheet -----------------------------------------

def test_multiple_sheets_and_hidden_sheet_detected(logged_in_client_a):
    data = _xlsx_bytes(
        [["מוצר", "מחיר"], ["X", "1"]],
        second_sheet_rows=[["Product", "Price"], ["Y", "2"]],
        hidden_sheet=True,
    )
    session_id, resp = _upload_and_analyze(logged_in_client_a, data)
    assert resp.status_code == 200
    sheets = resp.get_json()["analysis"]
    assert len(sheets) == 2
    assert sheets[0]["is_hidden"] is False
    assert sheets[1]["is_hidden"] is True

    session = logged_in_client_a.get(f"/api/imports/{session_id}").get_json()["session"]
    assert session["workbook_sheet_count"] == 2
    assert session["workbook_sheet_names"] == ["Sheet1", "Sheet2"]


# --- Duplicate headers -------------------------------------------------------

def test_duplicate_headers_detected_in_data_quality(logged_in_client_a):
    data = _xlsx_bytes([
        ["מוצר", 'לפני מע"מ', 'לפני מע"מ'],
        ["X", "10", "12"],
    ])
    session_id, resp = _upload_and_analyze(logged_in_client_a, data)
    sheet = resp.get_json()["analysis"][0]
    assert 'לפני מע"מ' in sheet["data_quality"]["duplicate_headers"]


# --- Unknown columns ---------------------------------------------------------

def test_unrecognizable_column_marked_unknown_not_guessed(logged_in_client_a):
    data = _xlsx_bytes([
        ["מוצר", "עמודה מסתורית"],
        ["X", "abc"],
        ["Y", "def"],
    ])
    session_id, resp = _upload_and_analyze(logged_in_client_a, data)
    sheet = resp.get_json()["analysis"][0]
    mystery = next(c for c in sheet["columns"] if c["header"] == "עמודה מסתורית")
    assert mystery["detected_type"] == "UNKNOWN"


# --- Empty workbook -----------------------------------------------------------

def test_empty_sheet_handled_gracefully(logged_in_client_a):
    """An upload with zero data rows fails at the staging layer (Phase
    3.1) before analysis ever runs — analysis only needs to not crash if
    called on such a session."""
    data = _xlsx_bytes([["מוצר", "מחיר"]])  # header only, staging marks this FAILED
    session_id = _upload(logged_in_client_a, data).get_json()["session"]["id"]
    resp = logged_in_client_a.post(f"/api/imports/{session_id}/analyze")
    assert resp.status_code == 200
    sheet = resp.get_json()["analysis"][0]
    assert sheet["row_count"] >= 0  # doesn't crash; header-only sheet still analyzable


def test_analyzing_truly_empty_sheet_does_not_crash(logged_in_client_a):
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    session_id = _upload(logged_in_client_a, buf.read()).get_json()["session"]["id"]
    resp = logged_in_client_a.post(f"/api/imports/{session_id}/analyze")
    assert resp.status_code == 200
    sheet = resp.get_json()["analysis"][0]
    assert sheet["row_count"] == 0
    assert sheet["detected_format"] == "UNKNOWN"


# --- Hebrew headers / general correctness -----------------------------------

def test_hebrew_headers_and_barcode_detection(logged_in_client_a):
    data = _xlsx_bytes([
        ["ברקוד", "מוצר", 'מק"ט', "כמות", "הערות"],
        ["7290000111222", "קולה 1.5", "COLA-15", "10", ""],
    ])
    session_id, resp = _upload_and_analyze(logged_in_client_a, data)
    sheet = resp.get_json()["analysis"][0]
    types = {c["header"]: c["detected_type"] for c in sheet["columns"]}
    assert types["ברקוד"] == "BARCODE"
    assert types["מוצר"] == "PRODUCT_NAME"
    assert types['מק"ט'] == "PRODUCT_CODE"
    assert types["כמות"] == "QUANTITY"
    assert types["הערות"] == "NOTES"


# --- Data quality: currency symbols, extra whitespace -----------------------

def test_currency_symbol_and_whitespace_detected(logged_in_client_a):
    data = _xlsx_bytes([
        ["מוצר", "מחיר"],
        ["  מוצר עם רווחים  ", "20.85 \u20aa"],
    ])
    session_id, resp = _upload_and_analyze(logged_in_client_a, data)
    sheet = resp.get_json()["analysis"][0]
    dq = sheet["data_quality"]
    assert "מחיר" in dq["currency_symbol_columns"]
    assert dq["extra_whitespace_cell_count"] >= 1


# --- Re-running analysis is idempotent (replaces, doesn't accumulate) -------

def test_rerunning_analysis_replaces_not_accumulates(logged_in_client_a):
    data = _xlsx_bytes([["מוצר", "מחיר"], ["X", "1"]])
    session_id = _upload(logged_in_client_a, data).get_json()["session"]["id"]
    logged_in_client_a.post(f"/api/imports/{session_id}/analyze")
    second = logged_in_client_a.post(f"/api/imports/{session_id}/analyze")
    assert second.status_code == 200
    assert len(second.get_json()["analysis"]) == 1  # not 2


# --- GET before analyze returns empty, not an error --------------------------

def test_get_analysis_before_analyzing_returns_empty_list(logged_in_client_a):
    data = _xlsx_bytes([["מוצר", "מחיר"], ["X", "1"]])
    session_id = _upload(logged_in_client_a, data).get_json()["session"]["id"]
    resp = logged_in_client_a.get(f"/api/imports/{session_id}/analysis")
    assert resp.status_code == 200
    assert resp.get_json()["analysis"] == []


# --- Permissions / tenant isolation ------------------------------------------

def test_analyze_requires_manager_role(client_a, tenant_a_admin):
    from test_security import _register_employee

    tenant_data, creds = tenant_a_admin
    slug = tenant_data["tenant"]["slug"]
    client_a.post("/api/auth/login", json=creds)
    data = _xlsx_bytes([["מוצר", "מחיר"], ["X", "1"]])
    session_id = _upload(client_a, data).get_json()["session"]["id"]

    _register_employee(client_a, slug)
    client_a.post("/api/auth/logout")
    client_a.post("/api/auth/login", json={"email": "worker@acme.test", "password": "Passw0rd1"})

    resp = client_a.post(f"/api/imports/{session_id}/analyze")
    assert resp.status_code == 403


def test_analysis_tenant_isolated(logged_in_client_a, logged_in_client_b):
    data = _xlsx_bytes([["מוצר", "מחיר"], ["X", "1"]])
    session_id = _upload(logged_in_client_a, data).get_json()["session"]["id"]
    logged_in_client_a.post(f"/api/imports/{session_id}/analyze")

    assert logged_in_client_b.post(f"/api/imports/{session_id}/analyze").status_code == 404
    assert logged_in_client_b.get(f"/api/imports/{session_id}/analysis").status_code == 404


# --- Hard guarantee: analysis never touches catalog tables -------------------

def test_analysis_never_touches_catalog_tables(logged_in_client_a):
    before_products = logged_in_client_a.get("/api/catalog/products").get_json()["products"]
    before_suppliers = logged_in_client_a.get("/api/catalog/suppliers").get_json()["suppliers"]

    data = _xlsx_bytes([
        [None, "גידרון", None, None, "עמית", "ווגשל"],
        ["מוצר", "יחידה", 'לפני מע"מ', "הנחה", 'לפני מע"מ', 'לפני מע"מ'],
        ["בורקס גבינה", "קילו", "20.85", "14.6", "12.8", "14.64"],
    ], merges=["B1:D1"])
    session_id, resp = _upload_and_analyze(logged_in_client_a, data)
    assert resp.status_code == 200
    assert resp.get_json()["analysis"][0]["detected_format"] == "WIDE"

    after_products = logged_in_client_a.get("/api/catalog/products").get_json()["products"]
    after_suppliers = logged_in_client_a.get("/api/catalog/suppliers").get_json()["suppliers"]
    assert after_products == before_products
    assert after_suppliers == before_suppliers


def test_audit_log_records_analysis(logged_in_client_a):
    data = _xlsx_bytes([["מוצר", "מחיר"], ["X", "1"]])
    session_id = _upload(logged_in_client_a, data).get_json()["session"]["id"]
    logged_in_client_a.post(f"/api/imports/{session_id}/analyze")
    logs = logged_in_client_a.get("/api/audit").get_json()["logs"]
    assert "import.analyzed" in [log["action"] for log in logs]


# --- Real uploaded files — not mock data only, per spec ---------------------

import pytest


@pytest.mark.skipif(not os.path.exists(REAL_FILE_PATH), reason="Real sample file not present in this environment")
def test_real_file_full_workbook_analysis(logged_in_client_a):
    """Analyzes the actual real ספקי מזון .xls file end to end — every
    sheet, no crashes, and the known-messy sheets produce sane results."""
    for name in ["גידרון", "עמית", "ווגשל"]:
        logged_in_client_a.post("/api/catalog/suppliers", json={"name": name})

    with open(REAL_FILE_PATH, "rb") as f:
        content = f.read()
    resp = _upload(logged_in_client_a, content, filename="real.xls", mime="application/vnd.ms-excel")
    session_id = resp.get_json()["session"]["id"]

    analyze_resp = logged_in_client_a.post(f"/api/imports/{session_id}/analyze")
    assert analyze_resp.status_code == 200
    sheets = analyze_resp.get_json()["analysis"]
    assert len(sheets) == 10  # this real workbook has 10 sheets

    session = logged_in_client_a.get(f"/api/imports/{session_id}").get_json()["session"]
    assert session["workbook_sheet_count"] == 10
    assert "גידרון" in session["workbook_sheet_names"]

    gidron = next(s for s in sheets if s["sheet_name"] == "גידרון")
    assert gidron["detected_format"] == "WIDE"
    assert gidron["header_tier_count"] == 2
    supplier_names = {s["header"] for s in gidron["detected_suppliers"]}
    assert {"גידרון", "עמית", "ווגשל"}.issubset(supplier_names)
    # The real sheet's used-range noise (~250 phantom columns) must be
    # flagged as empty, not silently miscounted as real data.
    assert len(gidron["data_quality"]["empty_columns"]) > 200

    yafora = next(s for s in sheets if s["sheet_name"] == "יפאורה")
    assert yafora["detected_format"] == "TALL"


def test_single_stray_value_in_phantom_column_does_not_pollute_orientation(logged_in_client_a):
    """Regression test for a real finding: a genuinely single-supplier
    (TALL) sheet with one leftover stray numeric value in an otherwise
    empty far-right "used range" column must still be detected as TALL —
    a sample size of 1 must not be enough to confidently type a column and
    tip orientation detection into MIXED."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["מוצר", "יחידה", "מחיר לפני מע\"מ", "הנחה", "מחיר אחרי הנחה"])
    for i in range(5):
        ws.append([f"מוצר {i}", "ארגז", 20 + i, 0.1, 18 + i])
    # A stray leftover value far outside the real data — simulates Excel's
    # inflated "used range" artifact confirmed against a real uploaded file.
    ws.cell(row=2, column=50, value=40.628)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    session_id, resp = _upload_and_analyze(logged_in_client_a, buf.read())
    sheet = resp.get_json()["analysis"][0]
    assert sheet["detected_format"] == "TALL", sheet["format_reason"]
